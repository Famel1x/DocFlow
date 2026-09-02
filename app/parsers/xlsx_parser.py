"""Парсер XLSX с поддержкой объединённых ячеек."""

import openpyxl
from typing import Any
from app.parsers.base import BaseParser


class XLSXParser(BaseParser):
    def parse(self) -> list[dict[str, Any]]:
        content = []
        wb = openpyxl.load_workbook(self.file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            content.append({"type": "text", "value": f"Лист: {sheet_name}"})

            max_row = sheet.max_row
            max_col = sheet.max_column
            if not max_row or not max_col:
                continue

            # Строим сетку значений
            grid: list[list[str]] = []
            for row in sheet.iter_rows(
                min_row=1, max_row=max_row, max_col=max_col, values_only=False
            ):
                row_data = []
                for cell in row:
                    val = cell.value
                    row_data.append(str(val) if val is not None else "")
                grid.append(row_data)

            # Заполняем объединённые ячейки: значение верхней-левой ячейки
            # копируется во все ячейки объединённого диапазона
            for merge_range in sheet.merged_cells.ranges:
                min_r = merge_range.min_row - 1  # 0-indexed
                min_c = merge_range.min_col - 1
                max_r = merge_range.max_row - 1
                max_c = merge_range.max_col - 1
                value = grid[min_r][min_c]
                for r in range(min_r, max_r + 1):
                    for c in range(min_c, max_c + 1):
                        grid[r][c] = value

            # Собираем метаданные об объединениях для фронтенда
            merges = []
            for merge_range in sheet.merged_cells.ranges:
                merges.append({
                    "min_row": merge_range.min_row - 1,
                    "min_col": merge_range.min_col - 1,
                    "max_row": merge_range.max_row - 1,
                    "max_col": merge_range.max_col - 1,
                })

            # Фильтруем полностью пустые строки
            table_data = [row for row in grid if any(cell.strip() for cell in row)]

            if table_data:
                content.append({
                    "type": "table",
                    "value": table_data,
                    "merges": merges,
                })

        return content
